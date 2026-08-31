# -*- coding: utf-8 -*-
"""
嗨回收抢单工具 - Web UI 服务端。
启动后浏览器访问 http://127.0.0.1:5868
"""

import os
import re
import sys
import threading
import time
import unicodedata
from datetime import datetime
from typing import Any, Dict, List, Optional

from flask import Flask, jsonify, render_template, request, send_file, session

from .api import HaihuishouAPI
from .grab_tool import GrabCondition, GrabOrderTool

# 成色 gradeId 与中文名对应（与接口 GetColorGradeLevel 一致）
COLOR_GRADE_ID_TO_NAME = {
    1001: "靓机",
    1002: "小花",
    1003: "大花",
    1004: "外爆",
    1005: "内爆",
}
# 成色中文名到 gradeId，用于入参兼容中文 key 时转成 ID key
COLOR_GRADE_NAME_TO_ID = {v: k for k, v in COLOR_GRADE_ID_TO_NAME.items()}
# 成色 ID 列表，报价时用 order.colorGrade 与条件中的 ID key 对比
COLOR_GRADE_IDS = (1001, 1002, 1003, 1004, 1005)

# 自动抢单：剩余时间少于此秒数的订单不抢（含已截止）
MIN_REMAINING_SECONDS = 30
_DEADLINE_KEYS = (
    "deadlineTime", "deadline_time", "expireAt", "expire_at",
    "endTime", "end_time", "deadline",
)
_REMAINING_KEYS = (
    "countdown", "remainSeconds", "remainingSeconds", "leftSeconds",
    "remain_seconds", "remaining_seconds", "left_seconds",
    "remainTime", "remainingTime",
)


def _pick_order_field(order: Dict[str, Any], keys) -> str:
    for k in keys:
        v = order.get(k)
        if v is not None and str(v).strip() != "":
            return str(v).strip()
    return ""


def _parse_deadline_epoch(v: Any) -> Optional[float]:
    """把截止时间解析成 epoch 秒；无法识别则返回 None。"""
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        n = float(v)
        if n > 1e12:
            return n / 1000.0
        if n > 1e9:
            return n
        return None
    s = str(v).strip()
    if not s:
        return None
    if s.isdigit() or (s.replace(".", "", 1).isdigit() and s.count(".") < 2):
        n = float(s)
        if n > 1e12:
            return n / 1000.0
        if n > 1e9:
            return n
        return None
    iso = s.replace(" ", "T", 1)
    try:
        return datetime.fromisoformat(iso).timestamp()
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s[:19], fmt).timestamp()
        except ValueError:
            continue
    return None


def _remaining_seconds(order: Dict[str, Any]) -> Optional[float]:
    raw = _pick_order_field(order, _DEADLINE_KEYS)
    if raw:
        ts = _parse_deadline_epoch(raw)
        if ts is not None:
            return ts - time.time()
    rem_raw = _pick_order_field(order, _REMAINING_KEYS)
    if rem_raw:
        try:
            n = float(str(rem_raw).strip())
            if n < 1e6:
                return n
        except (TypeError, ValueError):
            pass
    return None


def _deadline_too_soon(order: Dict[str, Any]) -> bool:
    """剩余时间已知且不足 MIN_REMAINING_SECONDS 时不抢。解析不到截止时间则不因此跳过。"""
    rem = _remaining_seconds(order)
    if rem is None:
        return False
    return rem < MIN_REMAINING_SECONDS


def _normalize_storage(s):
    """
    统一存储格式便于匹配。
    例：16G+256G| / 16+256G / 16G + 256G / 16GB+256GB -> 16g+256g
    """
    if s is None:
        return ""
    text = unicodedata.normalize("NFKC", str(s)).strip().lower()
    text = text.replace(" ", "")
    text = text.rstrip("|/\\,;，、")
    text = text.replace("gb", "g").replace("tb", "t")
    parts = re.findall(r"(\d+)\s*([gt])?", text)
    if not parts:
        return text
    normalized = []
    for num, unit in parts:
        normalized.append(num + (unit if unit else "g"))
    return "+".join(normalized)


def _normalize_color_name(s):
    """成色名规范化，避免全角/空格等导致匹配不到"""
    if s is None:
        return ""
    return unicodedata.normalize("NFKC", str(s).strip())


def _execute_task_fetch_order_list(cond, token, user_id):
    """
    抢单任务 - 查询待报价列表。
    查询时不带 conditions（机型/存储/成色价），只带分类、品牌、价格区间、厂商等 cond；
    conditions 在拿到列表后由 _execute_task_match_orders 做对比筛选。
    :param cond: GrabCondition（分类、品牌、价格、厂商等），不含抢单条件列表
    :param token: 登录 token
    :param user_id: 用户 ID
    :return: (lst, api)，lst 为订单列表，api 供后续抢单/报价使用
    """
    api = HaihuishouAPI()
    api.set_token(token, user_id)
    tool = GrabOrderTool(api=api)
    result = tool.step4_order_list(cond, page_index=1, user_id=user_id)
    lst = None
    if isinstance(result, dict):
        res_obj = result.get("result")
        if isinstance(res_obj, dict):
            lst = res_obj.get("orderList")
        if lst is None:
            lst = (
                result.get("list")
                or result.get("orderList")
                or result.get("results")
                or result.get("records")
                or result.get("rows")
                or result.get("items")
            )
        if lst is None and isinstance(result.get("data"), list):
            lst = result["data"]
        if lst is None and isinstance(result.get("data"), dict):
            inner = result["data"]
            lst = (inner.get("result") or {}).get("orderList") if isinstance(inner.get("result"), dict) else None
            lst = lst or inner.get("list") or inner.get("orderList") or inner.get("results") or []
    if not isinstance(lst, list):
        lst = []
    return lst, api


def _execute_task_match_orders(lst, conditions, use_conditions_list):
    """
    抢单任务 - 遍历订单列表，匹配条件并算出每条订单的报价金额。
    当成色为空或不存在时，仍加入匹配列表但 quote_amount 为 None，表示只抢单不报价。
    :param lst: 订单列表（来自查询列表接口）
    :param conditions: 归一化后的抢单条件
    :param use_conditions_list: 是否为新格式（含保底价/成色价）
    :return: ([(order, quote_amount), ...], skipped_short)
             quote_amount 为报价金额；skipped_short 为剩余不足 30 秒跳过的条数
    """
    def _quote_from_cond(order, cond):
        """
        按订单成色取报价：有成色时只用对应成色价（空则返回空，不抢单）；
        无成色时用保底价（空则返回空，不抢单）。成色价为空时不回退到保底价。
        """
        grade_id = order.get("colorGrade")
        if grade_id is not None and not (isinstance(grade_id, str) and not str(grade_id).strip()):
            try:
                key = str(int(grade_id))
                val = cond.get(key)
                if val is not None and str(val).strip():
                    return str(val).strip()
                return ""
            except (TypeError, ValueError):
                pass
        val = cond.get("floorPrice")
        if val is not None and str(val).strip():
            return str(val).strip()
        return ""

    def _find_condition_and_quote(order):
        """
        新格式：按机型+存储匹配条件，再取报价。
        无匹配或对应价格为空时返回 None（不抢单）；有价格则返回 (报价金额,)。
        """
        order_model = (order.get("modelName") or order.get("model") or order.get("goodsName") or "").strip().lower()
        order_storage = _normalize_storage(
            order.get("storageCapacity") or order.get("storage") or order.get("memory")
        )
        exact_match = None
        model_only_match = None
        for c in conditions:
            c_m = (c.get("modelName") or "").strip().lower()
            c_s = _normalize_storage(c.get("storage"))
            if c_m != order_model:
                continue
            if c_s:
                if c_s != order_storage:
                    continue
                exact_match = c
                break
            else:
                if model_only_match is None:
                    model_only_match = c
        chosen = exact_match or model_only_match
        if not chosen:
            return None
        quote = _quote_from_cond(order, chosen)
        if not quote:
            return None  # 对应价格为空，不抢单
        return (quote,)

    def _find_matching_condition(order_model, order_storage):
        """旧格式：按机型+存储匹配一条条件（条件无存储时只比机型）；无匹配返回 None。"""
        order_model = (order_model or "").strip().lower()
        order_storage = _normalize_storage(order_storage)
        for c in conditions:
            if (c.get("modelName") or "").strip().lower() != order_model:
                continue
            if c.get("storage"):
                if _normalize_storage(c.get("storage")) != order_storage:
                    continue
            return c
        return None

    matched = []
    skipped_short = 0
    for o in lst:
        if not isinstance(o, dict):
            continue
        record_id = o.get("recordId") or o.get("grabOrderId") or o.get("productId") or o.get("id")
        order_id = o.get("orderId") or o.get("orderNo") or o.get("orderSn")
        if record_id is None or order_id is None:
            continue
        if _deadline_too_soon(o):
            skipped_short += 1
            continue
        if use_conditions_list:
            res = _find_condition_and_quote(o)
            if res is None or not str(res[0] or "").strip():
                continue  # 未匹配或价格为空，不抢单
            matched.append((o, str(res[0]).strip()))
        else:
            order_model = (o.get("modelName") or o.get("model") or o.get("goodsName") or "").strip()
            order_storage = o.get("storageCapacity") or o.get("storage") or o.get("memory") or ""
            c = _find_matching_condition(order_model, order_storage)
            if not c:
                continue
            q = c.get("quoteAmount")
            if q and str(q).strip():
                matched.append((o, str(q).strip()))
            # 旧格式报价为空则不抢单
    return matched, skipped_short


def _execute_task_grab_and_quote(matched, api, user_id, remark):
    """
    抢单任务 - 对匹配到且有报价金额的订单执行抢单并提交报价。
    :param matched: [(order, quote_amount), ...]，quote_amount 须非空
    :param api: HaihuishouAPI 实例（已 set_token）
    :param user_id: 用户 ID
    :param remark: 报价备注
    :return: (grabbed, quoted, skipped_short, errors)
    """
    grabbed = 0
    quoted = 0
    skipped_short = 0
    errors = []
    for o, quote_for_submit in matched:
        record_id = o.get("recordId") or o.get("grabOrderId") or o.get("productId") or o.get("id")
        order_id = o.get("orderId") or o.get("orderNo") or o.get("orderSn")
        if quote_for_submit is None or not str(quote_for_submit).strip():
            continue
        # 匹配到实际抢单之间可能耗时，再次检查剩余时间
        if _deadline_too_soon(o):
            skipped_short += 1
            continue
        try:
            raw = api.grab_order(record_id=record_id, order_id=order_id, user_id=user_id)
            resp_data = raw.get("data") or {}
            sub_code = resp_data.get("subCode")
            if sub_code == 200:
                errors.append("recordId=%s 抢单失败: %s" % (record_id, (resp_data.get("subMessage") or "已被抢")))
                continue
            if sub_code != 100:
                errors.append("recordId=%s 抢单异常 subCode=%s" % (record_id, sub_code))
                continue
            grabbed += 1
            api.submit_quotation(
                record_id=int(record_id),
                order_id=int(order_id),
                actual_price=str(quote_for_submit).strip(),
                remark=remark,
                user_id=user_id,
            )
            quoted += 1
        except Exception as e:
            errors.append("recordId=%s: %s" % (record_id, str(e)))
    return grabbed, quoted, skipped_short, errors


# 打包成 exe 时模板在 sys._MEIPASS 下
_base_dir = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
_template_dir = os.path.join(_base_dir, "templates")
app = Flask(__name__, template_folder=_template_dir)
app.secret_key = os.environ.get("HAIHUISHOU_SECRET_KEY", "haihuishou-grab-dev-secret")
app.config["JSON_AS_ASCII"] = False

def _api_with_session() -> HaihuishouAPI:
    api = HaihuishouAPI()
    token = session.get("token")
    uid = session.get("user_id") or session.get("userId")
    if token:
        api.set_token(token, uid)
    return api


def _tool_with_session() -> GrabOrderTool:
    return GrabOrderTool(api=_api_with_session())


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json() or {}
    login_name = (data.get("loginName") or "").strip()
    login_pwd = data.get("loginPwd") or ""
    if not login_name or not login_pwd:
        return jsonify({"success": False, "message": "请填写手机号和密码"}), 400
    try:
        tool = _tool_with_session()
        info = tool.step1_login(login_name, login_pwd)
        session["token"] = info.get("token")
        uid = info.get("userId") or info.get("user_id")
        session["user_id"] = uid
        session["userId"] = uid
        return jsonify({"success": True, "data": info})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.pop("token", None)
    session.pop("user_id", None)
    session.pop("userId", None)
    return jsonify({"success": True})


@app.route("/api/status")
def api_status():
    token = session.get("token")
    uid = session.get("user_id") or session.get("userId")
    if token and uid:
        return jsonify({"loggedIn": True, "userId": uid, "token": token})
    return jsonify({"loggedIn": False, "userId": None})


@app.route("/api/user-info", methods=["GET"])
def api_user_info():
    """获取当前登录用户信息（姓名、手机、余额等）。"""
    api = _api_with_session()
    if not api.token or not api.user_id:
        return jsonify({"success": False, "message": "未登录"}), 401
    try:
        data = api.query_user_info()
        return jsonify({"success": True, "data": data})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/categories", methods=["GET"])
def api_categories():
    try:
        tool = _tool_with_session()
        data = tool.step2_manufacturer_and_categories()
        return jsonify({"success": True, "data": data})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/color-grade-level", methods=["GET"])
def api_color_grade_level():
    """获取手机成色列表（小花、大花、外爆、内爆等），用于前端筛选复选框。"""
    token = session.get("token")
    if not token:
        return jsonify({"success": False, "message": "请先登录"}), 401
    try:
        api = HaihuishouAPI()
        api.set_token(token, session.get("user_id") or session.get("userId"))
        lst = api.get_color_grade_level()
        return jsonify({"success": True, "data": lst})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/brands", methods=["GET"])
def api_brands():
    cat_id = request.args.get("catId", type=int)
    if cat_id is None:
        return jsonify({"success": False, "message": "缺少 catId"}), 400
    try:
        tool = _tool_with_session()
        brands = tool.step3_brands_by_category(cat_id)
        return jsonify({"success": True, "data": brands})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


def _parse_cell_number(val):
    """将单元格转为数字字符串，无效则返回空字符串。"""
    if val is None:
        return ""
    try:
        x = float(val)
        if x == int(x):
            return str(int(x))
        return str(x)
    except (TypeError, ValueError):
        return str(val).strip() if val is not None else ""


@app.route("/api/import-price-list", methods=["POST"])
def api_import_price_list():
    """上传价格列表 xlsx，解析 品牌、机型、内存、靓机、小花、大花、外爆、内爆、保底价 列，按品牌分组返回可批量创建任务的数据。"""
    if "file" not in request.files:
        return jsonify({"success": False, "message": "请选择要上传的文件"}), 400
    f = request.files["file"]
    fn = (f.filename or "").strip().lower()
    if not f.filename or not (fn.endswith(".xlsx") or fn.endswith(".xls")):
        return jsonify({"success": False, "message": "请上传 .xlsx 或 .xls 格式的 Excel 文件"}), 400
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        return jsonify({
            "success": False,
            "message": "服务端未安装 openpyxl，无法解析 Excel（打包时请用最新 haihuishou.spec 重新打包以打入 openpyxl）。详情: %s" % str(e),
        }), 500
    try:
        # 保存到临时文件再解析，避免上传流不可 seek 导致 read_only 模式失败
        import tempfile
        import os
        suffix = ".xlsx" if fn.endswith(".xlsx") else ".xls"
        tmp_path = None
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            f.save(tmp.name)
            tmp_path = tmp.name
        try:
            wb = load_workbook(tmp_path, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                return jsonify({"success": False, "message": "Excel 无有效工作表"}), 400
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        finally:
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass
        if not rows:
            return jsonify({"success": False, "message": "Excel 无数据"}), 400
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        col_brand = col_model = col_storage = None
        col_prices = {}  # 靓机、小花、大花、外爆、内爆、保底价
        price_keys = ("靓机", "小花", "大花", "外爆", "内爆", "保底价")
        # 内爆列可能表头为 内爆 或 内爆屏 等
        price_key_aliases = {"内爆": ("内爆", "内爆屏")}
        for i, h in enumerate(header):
            h_strip = (h or "").strip()
            if "品牌" in (h or "") or (h_strip and h_strip.lower() == "brand"):
                col_brand = i
            elif "机型" in (h or "") or (h_strip and h_strip.lower() in ("model", "型号")):
                col_model = i
            elif "内存" in (h or "") or "存储" in (h or "") or (h_strip and h_strip.lower() in ("storage", "memory")):
                col_storage = i
            else:
                for k in price_keys:
                    if col_prices.get(k) is not None:
                        continue
                    if k in (h or "") or h_strip == k:
                        col_prices[k] = i
                        break
                    aliases = price_key_aliases.get(k)
                    if aliases and h_strip in aliases:
                        col_prices[k] = i
                        break
        if col_brand is None or col_model is None:
            return jsonify({
                "success": False,
                "message": "Excel 需包含表头：品牌、机型（内存、靓机、小花、大花、外爆、内爆、保底价 可选）"
            }), 400
        by_brand = {}
        for row in rows[1:]:
            if not row:
                continue
            brand = (row[col_brand] if col_brand < len(row) else None)
            brand = str(brand).strip() if brand is not None else ""
            model = (row[col_model] if col_model < len(row) else None)
            model = str(model).strip() if model is not None else ""
            storage = ""
            if col_storage is not None and col_storage < len(row) and row[col_storage] is not None:
                storage = str(row[col_storage]).strip()
            if not brand or not model:
                continue
            # Excel 表头为中文，输出为成色 ID key（1001～1005）+ floorPrice；价格可为空
            cond = {"modelName": model, "storage": storage if storage else None}
            name_to_id = {"靓机": 1001, "小花": 1002, "大花": 1003, "外爆": 1004, "内爆": 1005}
            for name in price_keys:
                col_idx = col_prices.get(name)
                val = ""
                if col_idx is not None and col_idx < len(row):
                    val = _parse_cell_number(row[col_idx])
                if name == "保底价":
                    cond["floorPrice"] = val
                else:
                    gid = name_to_id.get(name)
                    if gid is not None:
                        cond[str(gid)] = val
            if brand not in by_brand:
                by_brand[brand] = []
            by_brand[brand].append(cond)
        result = [
            {"brandName": brand, "conditions": conds}
            for brand, conds in by_brand.items() if conds
        ]
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "message": "解析失败: %s" % str(e)}), 200


@app.route("/api/export-price-list", methods=["POST"])
def api_export_price_list():
    """根据前端传入的任务列表，生成与导入格式一致的 Excel（品牌、机型、内存、靓机、小花、大花、外爆、内爆、保底价），供下载。"""
    try:
        from openpyxl import Workbook
        import io
    except ImportError as e:
        return jsonify({
            "success": False,
            "message": "服务端未安装 openpyxl，无法生成 Excel（打包时请用最新 haihuishou.spec 重新打包以打入 openpyxl）。详情: %s" % str(e),
        }), 500
    body = request.get_json() or {}
    tasks = body.get("tasks")
    if not tasks or not isinstance(tasks, list):
        return jsonify({"success": False, "message": "请提供任务列表 tasks"}), 400
    # 表头与导入一致
    headers = ["品牌", "机型", "内存", "靓机", "小花", "大花", "外爆", "内爆", "保底价"]
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("价格列表", 0)
    ws.append(headers)
    for task in tasks:
        brand_name = (task.get("brandName") or "").strip() or (task.get("name") or "").replace("抢单", "").strip()
        conditions = task.get("conditions") or []
        for c in conditions:
            model = (c.get("modelName") or "").strip()
            storage = (c.get("storage") or "").strip()
            floor = (c.get("floorPrice") or "").strip()
            row = [
                brand_name,
                model,
                storage,
                (c.get("1001") or "").strip(),
                (c.get("1002") or "").strip(),
                (c.get("1003") or "").strip(),
                (c.get("1004") or "").strip(),
                (c.get("1005") or "").strip(),
                floor,
            ]
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="价格列表.xlsx",
    )


@app.route("/api/order-list", methods=["POST"])
def api_order_list():
    body = request.get_json() or {}
    # 优先用请求里的 token（headers）和 userId（body），没有则用 session
    token = request.headers.get("token") or session.get("token")
    user_id = body.get("userId") or session.get("user_id") or session.get("userId")
    if not token:
        return jsonify({"success": False, "message": "请先登录（缺少 token，需放在请求头）"}), 401
    if not user_id:
        return jsonify({"success": False, "message": "请先登录（缺少 userId，需放在请求体）"}), 401
    # categoryBrands: 电子产品+品牌 [{"key": "100001", "value": ["100007", "100011"]}]
    category_brands = body.get("categoryBrands") or []
    order_state = (body.get("orderState") or "10").strip()
    min_price = (body.get("minPrice") or "").strip() or None
    max_price = (body.get("maxPrice") or "").strip() or None
    sub_order_source_names = body.get("subOrderSourceNames") or []  # 厂商名称列表
    if isinstance(sub_order_source_names, str):
        sub_order_source_names = [x.strip() for x in sub_order_source_names.split(",") if x.strip()]
    _raw_grades = body.get("colorGradeIds") or []
    color_grade_ids = []
    if isinstance(_raw_grades, list):
        for x in _raw_grades:
            try:
                color_grade_ids.append(int(x))
            except (TypeError, ValueError):
                pass
    page = int(body.get("pageIndex", 1))
    page_size = int(body.get("pageSize", 20))
    cond = GrabCondition(
        category_brands=category_brands,
        order_state=order_state,
        min_price=min_price,
        max_price=max_price,
        sub_order_source_names=sub_order_source_names,
        color_grade_ids=color_grade_ids,
        page_size=page_size,
    )
    try:
        api = HaihuishouAPI()
        api.set_token(token, user_id)
        tool = GrabOrderTool(api=api)
        result = tool.step4_order_list(cond, page_index=page, user_id=user_id)
        # 出参：data.pageCount 为列表总数，data.result.orderList 为订单列表
        if isinstance(result, list):
            result = {"results": result, "totalCount": len(result)}
        elif isinstance(result, dict):
            lst = None
            res_obj = result.get("result")
            if isinstance(res_obj, dict):
                lst = res_obj.get("orderList")
            if lst is None:
                lst = (
                    result.get("list")
                    or result.get("orderList")
                    or result.get("results")
                    or result.get("records")
                    or result.get("rows")
                    or result.get("items")
                )
            if lst is None and isinstance(result.get("data"), list):
                lst = result["data"]
            if lst is None and isinstance(result.get("data"), dict):
                inner = result["data"]
                lst = inner.get("result", {}).get("orderList") if isinstance(inner.get("result"), dict) else None
                lst = lst or inner.get("list") or inner.get("orderList") or inner.get("results") or []
            if not isinstance(lst, list):
                lst = []
            total = result.get("pageCount") or result.get("totalCount")
            if total is None:
                total = len(lst)
            result = {"results": lst, "totalCount": total}
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/grab-order", methods=["POST"])
def api_grab_order():
    """先抢单，成功后再允许报价。body: recordId, orderId, userId；header: token。"""
    data = request.get_json() or {}
    token = request.headers.get("token") or session.get("token")
    user_id = data.get("userId") or session.get("user_id") or session.get("userId")
    if not token:
        return jsonify({"success": False, "message": "请先登录（缺少 token）"}), 401
    if not user_id:
        return jsonify({"success": False, "message": "请先登录（缺少 userId）"}), 401
    record_id = data.get("recordId")
    order_id = data.get("orderId")
    if record_id is None or record_id == "" or order_id is None or order_id == "":
        return jsonify({"success": False, "message": "缺少 recordId 或 orderId"}), 400
    try:
        api = HaihuishouAPI()
        api.set_token(token, user_id)
        raw = api.grab_order(record_id=record_id, order_id=order_id, user_id=user_id)
        resp_data = raw.get("data") or {}
        sub_code = resp_data.get("subCode")
        sub_message = (resp_data.get("subMessage") or "").strip()
        # 出参成功 subCode=100（抢单成功），失败 subCode=200（如已被其他报价师抢单）
        if sub_code == 200:
            return jsonify({"success": False, "message": sub_message or "抢单失败"}), 200
        if sub_code == 100:
            return jsonify({"success": True, "data": resp_data})
        return jsonify({"success": True, "data": resp_data})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/quote", methods=["POST"])
def api_quote():
    data = request.get_json() or {}
    token = request.headers.get("token") or session.get("token")
    user_id = data.get("userId") or session.get("user_id") or session.get("userId")
    if not token:
        return jsonify({"success": False, "message": "请先登录（缺少 token）"}), 401
    if not user_id:
        return jsonify({"success": False, "message": "请先登录（缺少 userId）"}), 401
    record_id = data.get("recordId")
    order_id = data.get("orderId")
    actual_price = data.get("actualPrice")  # 手动填写
    remark = (data.get("remark") or "").strip()  # 手动填写
    if record_id is None or order_id is None or actual_price is None or actual_price == "":
        return jsonify({"success": False, "message": "缺少 recordId / orderId / actualPrice（报价金额必填）"}), 400
    try:
        api = HaihuishouAPI()
        api.set_token(token, user_id)
        tool = GrabOrderTool(api=api)
        res = tool.step5_submit_quotation(
            record_id=int(record_id),
            order_id=int(order_id),
            actual_price=str(actual_price),
            quote_result=int(data.get("quoteResult", 1)),
            remark=remark,
            user_id=user_id,
        )
        return jsonify({"success": True, "data": res})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/execute-task", methods=["POST"])
def api_execute_task():
    """
    执行定时任务：查询待报价列表，按条件对象列表过滤，每条条件为 { quoteAmount, modelName, storage? }，订单匹配任一条件则用该条件的报价抢单。
    body: taskName, manufacturerNames[], categoryId, brandIds[], minPrice, maxPrice, conditions[]
    """
    data = request.get_json() or {}
    token = request.headers.get("token") or session.get("token")
    user_id = data.get("userId") or session.get("user_id") or session.get("userId")
    if not token:
        return jsonify({"success": False, "message": "请先登录（缺少 token）"}), 401
    if not user_id:
        return jsonify({"success": False, "message": "请先登录（缺少 userId）"}), 401
    task_name = (data.get("taskName") or "").strip()

    def task_err(msg):
        return jsonify({"success": False, "message": ("任务「%s」：%s" % (task_name, msg)) if task_name else msg}), 400

    manufacturer_names = data.get("manufacturerNames") or []
    if isinstance(manufacturer_names, str):
        manufacturer_names = [x.strip() for x in manufacturer_names.split(",") if x.strip()]
    category_id = (data.get("categoryId") or "").strip()
    brand_ids = data.get("brandIds") or []
    if isinstance(brand_ids, str):
        brand_ids = [str(x).strip() for x in brand_ids.split(",") if str(x).strip()]
    min_price = (data.get("minPrice") or "").strip() or None
    max_price = (data.get("maxPrice") or "").strip() or None
    conditions = data.get("conditions") or []
    if not conditions and (data.get("quoteAmount") or data.get("modelName")):
        conditions = [{
            "quoteAmount": str(data.get("quoteAmount", "")).strip(),
            "modelName": str(data.get("modelName", "")).strip(),
            "storage": (str(data.get("storage", "")).strip() or None),
        }]
    use_conditions_list = False
    if conditions and isinstance(conditions[0], dict) and ("floorPrice" in conditions[0] or "1001" in conditions[0]):
        use_conditions_list = True
        normalized = []
        seen_key = {}
        for idx, c in enumerate(conditions):
            if not isinstance(c, dict):
                continue
            m = (c.get("modelName") or "").strip()
            s = (c.get("storage") or "").strip() or ""
            if not m:
                continue
            floor_val = str(c.get("floorPrice") or "").strip()
            key = (m or "").lower() + "\x01" + (s or "").lower()
            if key in seen_key:
                first_no = seen_key[key]
                cur_no = idx + 1
                return task_err("序号 %d 与 序号 %d 机型+存储不可重复" % (first_no, cur_no))
            seen_key[key] = idx + 1
            row = {"modelName": m, "storage": s or None, "floorPrice": floor_val}
            for gid in COLOR_GRADE_IDS:
                key_id = str(gid)
                row[key_id] = str(c.get(key_id) or "").strip()
            normalized.append(row)
        if not normalized:
            return task_err("请至少添加一条条件（机型必填，价格可空）")
        conditions = normalized
    elif conditions:
        normalized = []
        for c in conditions:
            if not isinstance(c, dict):
                continue
            q = (c.get("quoteAmount") or "").strip()
            m = (c.get("modelName") or "").strip()
            s = (c.get("storage") or "").strip() or None
            if not m:
                continue
            # 报价可空：空则该条件不参与抢单
            normalized.append({"quoteAmount": q, "modelName": m, "storage": s})
        if not normalized:
            return task_err("请添加至少一条条件（机型必填）")
        conditions = normalized
    else:
        return task_err("请添加至少一条抢单条件")
    remark = task_name or "定时任务"
    category_brands = [{"key": category_id, "value": brand_ids}] if category_id else []
    # 查询列表：分类/品牌/价格/厂商 + 所有成色（colorGrade: [1001,1002,1003,1004,1005]）；conditions 仅用于后续对查询结果做匹配筛选
    cond = GrabCondition(
        category_brands=category_brands,
        order_state="10",
        min_price=min_price,
        max_price=max_price,
        sub_order_source_names=manufacturer_names,
        color_grade_ids=list(COLOR_GRADE_IDS),
        page_size=200,
    )
    try:
        lst, api = _execute_task_fetch_order_list(cond, token, user_id)
        matched, skipped_short_match = _execute_task_match_orders(lst, conditions, use_conditions_list)
        grabbed, quoted, skipped_short_grab, errors = _execute_task_grab_and_quote(matched, api, user_id, remark)
        return jsonify({
            "success": True,
            "data": {
                "grabbed": grabbed,
                "quoted": quoted,
                "total": len(lst),
                "skipped_short": skipped_short_match + skipped_short_grab,
                "errors": errors[:20],
            },
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/update-quote", methods=["POST"])
def api_update_quote():
    """已报价列表的修改报价，调用 hsdupdatequotation。body: recordId, orderId, actualPrice, remark, userId。"""
    data = request.get_json() or {}
    token = request.headers.get("token") or session.get("token")
    user_id = data.get("userId") or session.get("user_id") or session.get("userId")
    if not token:
        return jsonify({"success": False, "message": "请先登录（缺少 token）"}), 401
    if not user_id:
        return jsonify({"success": False, "message": "请先登录（缺少 userId）"}), 401
    record_id = data.get("recordId")
    order_id = data.get("orderId")
    actual_price = data.get("actualPrice")
    remark = (data.get("remark") or "").strip()
    if record_id is None or order_id is None or actual_price is None or actual_price == "":
        return jsonify({"success": False, "message": "缺少 recordId / orderId / actualPrice（报价金额必填）"}), 400
    try:
        api = HaihuishouAPI()
        api.set_token(token, user_id)
        res = api.update_quotation(
            record_id=record_id,
            order_id=order_id,
            actual_price=str(actual_price),
            remark=remark,
            user_id=user_id,
        )
        return jsonify({"success": True, "data": res})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/shutdown", methods=["POST"])
def api_shutdown():
    """试用结束：返回响应后退出进程，关闭工具。"""
    def _exit():
        import time
        time.sleep(1.5)
        os._exit(0)

    threading.Thread(target=_exit, daemon=False).start()
    return jsonify({"success": True, "message": "工具即将关闭"})


def main():
    host = os.environ.get("HAIHUISHOU_UI_HOST", "127.0.0.1")
    port = int(os.environ.get("HAIHUISHOU_UI_PORT", "5868"))
    debug = not getattr(sys, "frozen", False)
    if not debug:
        import webbrowser
        import threading
        def _open_browser():
            import time
            time.sleep(1.2)
            webbrowser.open(f"http://{host}:{port}")
        threading.Thread(target=_open_browser, daemon=True).start()
    app.run(host=host, port=port, debug=debug)


if __name__ == "__main__":
    main()
