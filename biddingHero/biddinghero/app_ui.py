# -*- coding: utf-8 -*-
"""
竞价侠 - Web UI 服务端。
启动后浏览器访问 http://127.0.0.1:5060
"""

import os
import sys
import threading
from typing import Any, Dict, List

from flask import Flask, jsonify, render_template, request, send_file, session

from .api import BiddingHeroAPI
from .grab_task import (
    CONDITION_PRICE_KEYS,
    GRADE_KEYS,
    GRADE_NAMES,
    GrabCondition,
    execute_task,
    normalize_conditions,
)


_base_dir = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
_template_dir = os.path.join(_base_dir, "templates")
app = Flask(__name__, template_folder=_template_dir)
app.secret_key = os.environ.get("BIDDINGHERO_SECRET_KEY", "biddinghero-dev-secret")
app.config["JSON_AS_ASCII"] = False


def _api_with_session() -> BiddingHeroAPI:
    api = BiddingHeroAPI()
    token = session.get("token")
    uid = session.get("user_id") or session.get("userId")
    if token:
        api.set_token(token, uid)
    return api


def _extract_order_id(item: Any) -> Any:
    """从报价中列表 item 里取商品 id；item 可能就是 id 字符串或数字。"""
    if isinstance(item, dict):
        for key in ("orderId", "order_id", "id"):
            v = item.get(key)
            if v is not None:
                return v
        return None
    return item


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json() or {}
    phone = (data.get("phone") or data.get("loginName") or "").strip()
    password = data.get("password") or data.get("loginPwd") or ""
    if not phone or not password:
        return jsonify({"success": False, "message": "请填写手机号和密码"}), 400
    try:
        api = BiddingHeroAPI()
        info = api.login(phone, password)
        session["token"] = api.token
        uid = api.user_id
        session["user_id"] = uid
        session["userId"] = uid
        return jsonify({"success": True, "data": info})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.pop("token", None)
    session.pop("user_id", None)
    session.pop("userId", None)
    return jsonify({"success": True})


@app.route("/api/status")
def api_status():
    token = session.get("token")
    uid = session.get("user_id") or session.get("userId")
    if token and uid:
        return jsonify({"loggedIn": True, "userId": uid, "token": token})
    return jsonify({"loggedIn": False, "userId": None})


@app.route("/api/user-info", methods=["GET"])
def api_user_info():
    """获取当前登录用户信息。"""
    api = _api_with_session()
    if not api.token:
        return jsonify({"success": False, "message": "未登录"}), 401
    try:
        data = api.get_my_info()
        return jsonify({"success": True, "data": data})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/order-list", methods=["GET", "POST"])
def api_order_list():
    """待竞价列表。"""
    body = request.get_json(silent=True) or {}
    api = _api_with_session()
    if not api.token:
        return jsonify({"success": False, "message": "请先登录"}), 401
    page_index = int(body.get("pageIndex") or request.args.get("pageIndex") or 1)
    page_size = int(body.get("pageSize") or request.args.get("pageSize") or 3000)
    try:
        data = api.get_auction_list(page_index=page_index, page_size=page_size)
        # 统一结构：results / totalCount
        results = data.get("results") if isinstance(data, dict) else None
        total = data.get("count") if isinstance(data, dict) else None
        if results is None and isinstance(data, list):
            results = data
        if results is None:
            results = []
        if total is None:
            total = len(results)
        return jsonify({"success": True, "data": {"results": results, "totalCount": total}})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/grab-order-list", methods=["GET"])
def api_grab_order_list():
    """报价中列表：查看详情后商品会进入这个列表。"""
    api = _api_with_session()
    if not api.token:
        return jsonify({"success": False, "message": "请先登录"}), 401
    try:
        lst = api.get_grab_order_list()
        # 提取纯 id 列表，供前端做对比
        ids: List[Any] = []
        for it in lst or []:
            oid = _extract_order_id(it)
            if oid is not None:
                ids.append(oid)
        return jsonify({"success": True, "data": {"results": lst or [], "ids": ids}})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/my-bids", methods=["GET"])
def api_my_bids():
    """已报价列表。"""
    api = _api_with_session()
    if not api.token:
        return jsonify({"success": False, "message": "请先登录"}), 401
    page_index = int(request.args.get("pageIndex", 1))
    page_size = int(request.args.get("pageSize", 20))
    status = (request.args.get("status") or "").strip() or None
    created_at_after = (request.args.get("created_at_after") or "").strip() or None
    created_at_before = (request.args.get("created_at_before") or "").strip() or None
    try:
        data = api.get_my_bids(
            page_index=page_index,
            page_size=page_size,
            status=status,
            created_at_after=created_at_after,
            created_at_before=created_at_before,
        )
        results = data.get("results") if isinstance(data, dict) else None
        total = data.get("count") if isinstance(data, dict) else None
        if results is None:
            results = []
        if total is None:
            total = len(results)
        return jsonify({"success": True, "data": {"results": results, "totalCount": total}})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/order-detail", methods=["GET"])
def api_order_detail():
    """商品详情。看完详情后此商品即进入报价中列表。"""
    api = _api_with_session()
    if not api.token:
        return jsonify({"success": False, "message": "请先登录"}), 401
    order_id = request.args.get("orderId")
    if not order_id:
        return jsonify({"success": False, "message": "缺少 orderId"}), 400
    try:
        data = api.get_order_detail(order_id)
        report = api.get_order_report_safe(order_id)
        # 再拉一次报价中列表，返回给前端做同步
        grab_ids: List[Any] = []
        try:
            grab_list = api.get_grab_order_list()
            for it in grab_list or []:
                oid = _extract_order_id(it)
                if oid is not None:
                    grab_ids.append(oid)
        except Exception:
            grab_list = []
        return jsonify({
            "success": True,
            "data": {
                "detail": data,
                "report": report,
                "grabList": grab_list,
                "grabIds": grab_ids,
            },
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/place-bid", methods=["POST"])
def api_place_bid():
    """出价接口。前端可直接从列表出价。"""
    data = request.get_json() or {}
    api = _api_with_session()
    if not api.token:
        return jsonify({"success": False, "message": "请先登录"}), 401
    order_id = data.get("orderId")
    bid_amount = data.get("bidAmount") or data.get("actualPrice")
    if not order_id or bid_amount in (None, ""):
        return jsonify({"success": False, "message": "缺少 orderId / bidAmount"}), 400
    try:
        res = api.place_bid(order_id=order_id, bid_amount=bid_amount)
        return jsonify({"success": True, "data": res})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/cancel-grab-order", methods=["POST"])
def api_cancel_grab_order():
    """取消抢单（把商品从报价中列表移除）。"""
    data = request.get_json() or {}
    api = _api_with_session()
    if not api.token:
        return jsonify({"success": False, "message": "请先登录"}), 401
    order_id = data.get("orderId")
    if not order_id:
        return jsonify({"success": False, "message": "缺少 orderId"}), 400
    try:
        res = api.cancel_grab_order(order_id=order_id)
        return jsonify({"success": True, "data": res})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/reference-price", methods=["GET"])
def api_reference_price():
    """根据订单编号查询市场价格参考。"""
    api = _api_with_session()
    if not api.token:
        return jsonify({"success": False, "message": "请先登录"}), 401
    order_number = (request.args.get("order_number") or request.args.get("orderNumber") or "").strip()
    if not order_number:
        return jsonify({"success": False, "message": "缺少 order_number"}), 400
    try:
        data = api.get_reference_price(order_number)
        return jsonify({"success": True, "data": data})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/order-report", methods=["GET"])
def api_order_report():
    """验机报告：包含内存/颜色/购买渠道等 inspection_items。"""
    api = _api_with_session()
    if not api.token:
        return jsonify({"success": False, "message": "请先登录"}), 401
    order_id = request.args.get("orderId")
    if not order_id:
        return jsonify({"success": False, "message": "缺少 orderId"}), 400
    try:
        data = api.get_order_report(order_id)
        return jsonify({"success": True, "data": data})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


@app.route("/api/execute-task", methods=["POST"])
def api_execute_task():
    """执行自动抢单任务。body: taskName, categoryNames[], brandNames[], channelNames[], minPrice, maxPrice, maxAmount, conditions[]。"""
    api = _api_with_session()
    if not api.token:
        return jsonify({"success": False, "message": "请先登录"}), 401
    data = request.get_json() or {}
    task_name = (data.get("taskName") or "").strip()

    def task_err(msg):
        return jsonify({"success": False, "message": ("任务「%s」：%s" % (task_name, msg)) if task_name else msg}), 400

    category_names = data.get("categoryNames") or []
    brand_names = data.get("brandNames") or []
    channel_names = data.get("channelNames") or []
    for name, val in (("categoryNames", category_names), ("brandNames", brand_names), ("channelNames", channel_names)):
        if isinstance(val, str):
            if name == "categoryNames":
                category_names = [x.strip() for x in val.split(",") if x.strip()]
            elif name == "brandNames":
                brand_names = [x.strip() for x in val.split(",") if x.strip()]
            else:
                channel_names = [x.strip() for x in val.split(",") if x.strip()]
    min_price = (str(data.get("minPrice") or "").strip()) or None
    max_price = (str(data.get("maxPrice") or "").strip()) or None
    max_amount_raw = str(data.get("maxAmount") or "").strip()
    try:
        max_amount = float(max_amount_raw) if max_amount_raw else None
    except ValueError:
        return task_err("最大金额格式错误")
    conditions_raw = data.get("conditions") or []
    conditions, err = normalize_conditions(conditions_raw)
    if err:
        return task_err(err)
    cond = GrabCondition(
        category_names=category_names,
        brand_names=brand_names,
        channel_names=channel_names,
        min_price=min_price,
        max_price=max_price,
        max_amount=max_amount,
        conditions=conditions,
    )
    try:
        result = execute_task(api, cond)
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 200


def _parse_cell_number(val):
    if val is None:
        return ""
    try:
        x = float(val)
        if x == int(x):
            return str(int(x))
        return str(x)
    except (TypeError, ValueError):
        return str(val).strip() if val is not None else ""


@app.route("/api/import-price-list", methods=["POST"])
def api_import_price_list():
    """导入价格 Excel，列：品牌、机型、内存、全新、靓机、小花、大花、外爆、内爆、功能异常、维修更换、无法开机、保底价。"""
    if "file" not in request.files:
        return jsonify({"success": False, "message": "请选择要上传的文件"}), 400
    f = request.files["file"]
    fn = (f.filename or "").strip().lower()
    if not f.filename or not (fn.endswith(".xlsx") or fn.endswith(".xls")):
        return jsonify({"success": False, "message": "请上传 .xlsx 或 .xls 格式的 Excel 文件"}), 400
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        return jsonify({
            "success": False,
            "message": "服务端未安装 openpyxl，无法解析 Excel。详情: %s" % str(e),
        }), 500
    try:
        import tempfile
        suffix = ".xlsx" if fn.endswith(".xlsx") else ".xls"
        tmp_path = None
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            f.save(tmp.name)
            tmp_path = tmp.name
        try:
            wb = load_workbook(tmp_path, read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                return jsonify({"success": False, "message": "Excel 无有效工作表"}), 400
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        finally:
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass
        if not rows:
            return jsonify({"success": False, "message": "Excel 无数据"}), 400
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        col_brand = col_model = col_storage = None
        col_prices: Dict[str, int] = {}
        price_keys = list(GRADE_NAMES) + ["保底价"]
        for i, h in enumerate(header):
            h_strip = (h or "").strip()
            if "品牌" in (h or "") or h_strip.lower() == "brand":
                col_brand = i
            elif "机型" in (h or "") or h_strip.lower() in ("model", "型号"):
                col_model = i
            elif "内存" in (h or "") or "存储" in (h or "") or h_strip.lower() in ("storage", "memory"):
                col_storage = i
            else:
                for k in price_keys:
                    if col_prices.get(k) is not None:
                        continue
                    if k in (h or "") or h_strip == k:
                        col_prices[k] = i
                        break
        if col_brand is None or col_model is None:
            return jsonify({
                "success": False,
                "message": "Excel 需包含表头：品牌、机型（内存与 9 档成色 + 保底价 可选）"
            }), 400
        by_brand: Dict[str, List[Dict[str, Any]]] = {}
        for row in rows[1:]:
            if not row:
                continue
            brand = row[col_brand] if col_brand < len(row) else None
            brand = str(brand).strip() if brand is not None else ""
            model = row[col_model] if col_model < len(row) else None
            model = str(model).strip() if model is not None else ""
            storage = ""
            if col_storage is not None and col_storage < len(row) and row[col_storage] is not None:
                storage = str(row[col_storage]).strip()
            if not brand or not model:
                continue
            cond: Dict[str, Any] = {"modelName": model, "storage": storage if storage else None}
            for name in price_keys:
                col_idx = col_prices.get(name)
                val = ""
                if col_idx is not None and col_idx < len(row):
                    val = _parse_cell_number(row[col_idx])
                if name == "保底价":
                    cond["floorPrice"] = val
                else:
                    cond["q_" + name] = val
            by_brand.setdefault(brand, []).append(cond)
        result = [{"brandName": b, "conditions": cs} for b, cs in by_brand.items() if cs]
        return jsonify({"success": True, "data": result})
    except Exception as e:
        return jsonify({"success": False, "message": "解析失败: %s" % str(e)}), 200


@app.route("/api/export-price-list", methods=["POST"])
def api_export_price_list():
    try:
        from openpyxl import Workbook
        import io
    except ImportError as e:
        return jsonify({"success": False, "message": "服务端未安装 openpyxl。详情: %s" % str(e)}), 500
    body = request.get_json() or {}
    tasks = body.get("tasks")
    if not tasks or not isinstance(tasks, list):
        return jsonify({"success": False, "message": "请提供任务列表 tasks"}), 400
    headers = ["品牌", "机型", "内存"] + list(GRADE_NAMES) + ["保底价"]
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("价格列表", 0)
    ws.append(headers)
    for task in tasks:
        brand_name = (task.get("brandName") or "").strip() or (task.get("name") or "").replace("抢单", "").strip()
        for c in (task.get("conditions") or []):
            row = [
                brand_name,
                (c.get("modelName") or "").strip(),
                (c.get("storage") or "").strip(),
            ]
            for name in GRADE_NAMES:
                row.append((c.get("q_" + name) or "").strip())
            row.append((c.get("floorPrice") or "").strip())
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="价格列表.xlsx",
    )


@app.route("/api/shutdown", methods=["POST"])
def api_shutdown():
    def _exit():
        import time
        time.sleep(1.5)
        os._exit(0)

    threading.Thread(target=_exit, daemon=False).start()
    return jsonify({"success": True, "message": "工具即将关闭"})


def main():
    host = os.environ.get("BIDDINGHERO_UI_HOST", "127.0.0.1")
    port = int(os.environ.get("BIDDINGHERO_UI_PORT", "5070"))
    debug = not getattr(sys, "frozen", False)
    if not debug:
        import webbrowser
        def _open_browser():
            import time
            time.sleep(1.2)
            webbrowser.open("http://{}:{}".format(host, port))
        threading.Thread(target=_open_browser, daemon=True).start()
    app.run(host=host, port=port, debug=debug)


if __name__ == "__main__":
    main()
